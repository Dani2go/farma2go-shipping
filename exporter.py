"""
exporter.py — Generate Excel P&L reports
"""

import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = 'Arial'
C_HDR = '1F4E79'; C_SUB = '2E75B6'
C_GRN = 'E2EFDA'; C_RED = 'FFDCDC'; C_GREY = 'F5F5F5'; C_YEL = 'FFF2CC'
EUR2 = '#,##0.00€;[Red]-#,##0.00€'; PCT = '0.0%'; INT_F = '#,##0'

CTY_COLORS = {
    'España':'1F4E79','Portugal':'375623','Francia':'833C00',
    'Italia':'843C0C','Alemania':'4472C4','Reino Unido':'7030A0',
    'DEFAULT':'5A5A5A',
}

def _fill(h): return PatternFill('solid', start_color=h, fgColor=h)
def _bdr():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _dv(ws, r, ci, v, fmt=None, bg=None, bold=False, fc=None, size=10):
    c = ws.cell(r, ci, v)
    c.font = Font(name=FONT, size=size, bold=bold, color=fc or '000000')
    c.border = _bdr()
    if bg: c.fill = _fill(bg)
    if fmt: c.number_format = fmt
    c.alignment = Alignment(
        horizontal='right' if isinstance(v, (int, float)) else 'left',
        vertical='center')

def _mc(ws, r, ci, v, fmt=EUR2, bold=False, size=10):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        c = ws.cell(r, ci, '—')
        c.font = Font(name=FONT, size=size, color='CCCCCC')
        c.border = _bdr()
        c.alignment = Alignment(horizontal='center', vertical='center')
        return
    bg = C_GRN if v >= 0 else C_RED
    fc = '006400' if v >= 0 else 'CC0000'
    _dv(ws, r, ci, float(v), fmt=fmt, bg=bg, bold=bold, fc=fc, size=size)

def _hrow(ws, r, vals, bg, h=20):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(r, ci, v)
        c.font = Font(name=FONT, bold=True, size=10, color='FFFFFF')
        c.fill = _fill(bg); c.border = _bdr()
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = h

def _title(ws, r, nc, text, bg=None):
    ws.merge_cells(f'A{r}:{get_column_letter(nc)}{r}')
    ws[f'A{r}'].value = text
    ws[f'A{r}'].font = Font(name=FONT, bold=True, size=13, color='FFFFFF')
    ws[f'A{r}'].fill = _fill(bg or C_HDR)
    ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 28

def _widths(ws, wlist):
    for i, w in enumerate(wlist, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_pnl_excel(pnl_data: dict, month_label: str = '') -> bytes:
    """Generate full P&L Excel from engine output."""
    wb = Workbook(); wb.remove(wb.active)
    title_suffix = f' — {month_label}' if month_label else ''

    # ── SHEET 1: RESUMEN ─────────────────────────────────────────
    ws1 = wb.create_sheet('📊 Resumen P&L')
    _title(ws1, 1, 6, f'P&L FARMA2GO{title_suffix}')
    r = 3

    # Summary KPIs
    _hrow(ws1, r, ['Concepto', 'Valor', '', '', '', ''], C_HDR, h=22); r += 1

    kpis = []
    if 'pnl_by_country' in pnl_data:
        rows = pnl_data['pnl_by_country']
        total_venta = sum(x.get('venta', 0) for x in rows)
        total_cogs  = sum(x.get('cogs', 0) for x in rows)
        total_mg    = sum(x.get('mg_prod', 0) for x in rows)
        total_ing_e = sum(x.get('ing_envio', 0) for x in rows)
        total_cost_e= sum(x.get('cost_envio', 0) for x in rows)
        total_mf    = sum(x.get('mg_final', 0) for x in rows)
        kpis = [
            ('Venta bruta productos', total_venta, EUR2, False),
            ('COGS (coste mercancía)', total_cogs, EUR2, False),
            ('Margen bruto producto', total_mg, EUR2, True),
            ('% Margen bruto', total_mg / total_venta if total_venta else 0, PCT, True),
            ('Ingreso envío cobrado', total_ing_e, EUR2, False),
            ('Coste envío carrier', total_cost_e, EUR2, False),
            ('Margen envío', total_ing_e - total_cost_e, EUR2, True),
            ('MARGEN FINAL', total_mf, EUR2, True),
            ('% Margen final', total_mf / (total_venta + total_ing_e) if (total_venta + total_ing_e) else 0, PCT, True),
        ]
    elif 'shipping' in pnl_data:
        rows = pnl_data['shipping']
        total_cost  = sum(x.get('coste_total', 0) for x in rows)
        total_ing   = sum(x.get('ingreso_total', 0) for x in rows)
        total_mg    = sum(x.get('margen_envio', 0) for x in rows)
        kpis = [
            ('Coste envío (carrier)', total_cost, EUR2, False),
            ('Ingreso envío cobrado', total_ing, EUR2, False),
            ('Margen envío', total_mg, EUR2, True),
        ]

    if 'total_ads' in pnl_data:
        ads_cost = pnl_data['total_ads']
        if kpis:
            base = kpis[-3][1] if len(kpis) >= 3 else 0
            net = (kpis[-2][1] if len(kpis) >= 2 else 0) - ads_cost
            kpis += [
                ('', None, None, False),
                ('Gasto Google Ads', ads_cost, EUR2, False),
                ('MARGEN POST-ADS', net, EUR2, True),
            ]

    for label, val, fmt, is_m in kpis:
        if val is None:
            ws1.row_dimensions[r].height = 8; r += 1; continue
        ws1.cell(r, 1, label).font = Font(name=FONT, bold=True, size=10)
        ws1.cell(r, 1).border = _bdr(); ws1.cell(r, 1).fill = _fill(C_GREY)
        ws1.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
        v = float(val) if isinstance(val, float) else val
        _mc(ws1, r, 2, v, fmt=fmt, bold=True) if is_m else _dv(ws1, r, 2, v, fmt=fmt, bold=True)
        for ci in range(3, 7): ws1.cell(r, ci).border = _bdr()
        r += 1

    _widths(ws1, [30, 16, 16, 16, 16, 16])

    # ── SHEET 2: POR PAÍS ────────────────────────────────────────
    if 'pnl_by_country' in pnl_data:
        ws2 = wb.create_sheet('🌍 Por País')
        rows = pnl_data['pnl_by_country']
        _title(ws2, 1, 8, f'P&L POR PAÍS{title_suffix}')
        r = 3
        _hrow(ws2, r, ['País','Pedidos','Venta €','COGS €','Mg Prod €','Ing Envío €','Coste Envío €','Margen Final €'], C_HDR); r += 1

        by_country = {}
        for row in rows:
            c = row.get('country', 'Desconocido')
            if c not in by_country: by_country[c] = {k: 0 for k in ['n_pedidos','venta','cogs','mg_prod','ing_envio','cost_envio','mg_final']}
            for k in by_country[c]: by_country[c][k] += row.get(k, 0)

        for country, d in sorted(by_country.items(), key=lambda x: -x[1]['venta']):
            col = CTY_COLORS.get(country, CTY_COLORS['DEFAULT'])
            ws2.cell(r, 1, country).font = Font(name=FONT, bold=True, size=10, color='FFFFFF')
            ws2.cell(r, 1).fill = _fill(col); ws2.cell(r, 1).border = _bdr()
            ws2.cell(r, 1).alignment = Alignment(horizontal='left', vertical='center')
            _dv(ws2, r, 2, int(d['n_pedidos']), fmt=INT_F)
            _dv(ws2, r, 3, float(d['venta']), fmt=EUR2)
            _dv(ws2, r, 4, float(d['cogs']), fmt=EUR2)
            _mc(ws2, r, 5, float(d['mg_prod']))
            _dv(ws2, r, 6, float(d['ing_envio']), fmt=EUR2)
            _dv(ws2, r, 7, float(d['cost_envio']), fmt=EUR2, bg=C_RED if d['cost_envio'] > d['ing_envio'] else None)
            _mc(ws2, r, 8, float(d['mg_final']), bold=True)
            r += 1

        _widths(ws2, [16, 10, 16, 16, 14, 14, 14, 16])

    # ── SHEET 3: POR CARRIER ─────────────────────────────────────
    if 'shipping' in pnl_data:
        ws3 = wb.create_sheet('🚚 Por Carrier')
        rows = pnl_data['shipping']
        _title(ws3, 1, 7, f'MARGEN ENVÍOS POR CARRIER{title_suffix}')
        r = 3
        _hrow(ws3, r, ['Carrier','País','Envíos','Coste €','Ingreso €','Margen €','€/envío'], C_HDR); r += 1

        for row in sorted(rows, key=lambda x: x.get('margen_envio', 0)):
            _dv(ws3, r, 1, str(row.get('carrier', '')), bold=True)
            _dv(ws3, r, 2, str(row.get('country', '')))
            _dv(ws3, r, 3, int(row.get('n_envios', 0)), fmt=INT_F)
            _dv(ws3, r, 4, float(row.get('coste_total', 0)), fmt=EUR2)
            _dv(ws3, r, 5, float(row.get('ingreso_total', 0)), fmt=EUR2)
            _mc(ws3, r, 6, float(row.get('margen_envio', 0)))
            n = row.get('n_envios', 1) or 1
            _mc(ws3, r, 7, float(row.get('margen_envio', 0)) / n, fmt='#,##0.00€')
            r += 1

        _widths(ws3, [12, 14, 10, 14, 14, 14, 12])

    # ── SHEET 4: ALERTAS RECLAMACIÓN ─────────────────────────────
    if 'alerts' in pnl_data and pnl_data['alerts']:
        ws4 = wb.create_sheet('⚠️ Reclamaciones')
        alerts = pnl_data['alerts']
        _title(ws4, 1, 8, f'ENVÍOS PARA RECLAMAR — Pérdida > 8€ en envíos PAGADOS{title_suffix}', bg='C00000')
        r = 3

        # Summary note
        ws4.merge_cells(f'A{r}:H{r}')
        ws4[f'A{r}'].value = (f"⚠️  {len(alerts)} envíos con pérdida anormal "
                               f"| Pérdida total: {pnl_data.get('alert_total_loss', 0):,.2f}€")
        ws4[f'A{r}'].font = Font(name=FONT, bold=True, size=11, color='C00000')
        ws4[f'A{r}'].fill = _fill('FFF2CC')
        ws4[f'A{r}'].alignment = Alignment(horizontal='left', vertical='center')
        ws4.row_dimensions[r].height = 24; r += 1

        _hrow(ws4, r, ['Ref Pedido','Carrier','País','Peso kg','Cobrado €','Coste €','Pérdida €','Ratio coste/cobrado'], C_HDR, h=28); r += 1
        ws4.freeze_panes = 'A4'

        for alert in sorted(alerts, key=lambda x: x.get('margin', 0)):
            ratio = abs(float(alert.get('cost_eur', 0))) / max(float(alert.get('precio_envio', 1)), 0.01)
            _dv(ws4, r, 1, str(alert.get('ref', '')), bold=True)
            _dv(ws4, r, 2, str(alert.get('carrier', '')))
            _dv(ws4, r, 3, str(alert.get('country', '')))
            _dv(ws4, r, 4, float(alert.get('weight_kg', 0)), fmt='0.00 kg')
            _dv(ws4, r, 5, float(alert.get('precio_envio', 0)), fmt=EUR2)
            _dv(ws4, r, 6, float(alert.get('cost_eur', 0)), fmt=EUR2, bg=C_RED, fc='CC0000')
            _mc(ws4, r, 7, float(alert.get('margin', 0)), bold=True)
            c_ratio = ws4.cell(r, 8, round(ratio, 1))
            c_ratio.number_format = '0.0x'
            c_ratio.font = Font(name=FONT, size=10, bold=True, color='CC0000' if ratio >= 3 else '7F6000')
            c_ratio.border = _bdr()
            c_ratio.alignment = Alignment(horizontal='right', vertical='center')
            r += 1

        _widths(ws4, [14, 12, 14, 10, 14, 14, 14, 18])

    # ── SHEET 5: GOOGLE ADS ──────────────────────────────────────
    if 'ads' in pnl_data:
        ws5 = wb.create_sheet('📣 Google Ads')
        ads = pnl_data['ads']
        _title(ws5, 1, 7, f'GOOGLE ADS{title_suffix}')
        r = 3
        _hrow(ws5, r, ['País','Mes','Gasto €','Conversiones','Valor Conv €','ROAS','CPA €'], C_HDR); r += 1

        for row in sorted(ads, key=lambda x: (x.get('ym',''), -x.get('gasto_ads', 0))):
            _dv(ws5, r, 1, str(row.get('pais', '')), bold=True)
            _dv(ws5, r, 2, str(row.get('ym', '')))
            _dv(ws5, r, 3, float(row.get('gasto_ads', 0)), fmt=EUR2)
            _dv(ws5, r, 4, float(row.get('conversiones', 0)), fmt=INT_F)
            _dv(ws5, r, 5, float(row.get('valor_conv', 0)), fmt=EUR2)
            roas = float(row.get('roas', 0) or 0)
            c_r = ws5.cell(r, 6, round(roas, 1))
            c_r.number_format = '0.0'
            c_r.font = Font(name=FONT, size=10, color='006400' if roas >= 6 else ('7F6000' if roas >= 3 else 'CC0000'))
            c_r.border = _bdr(); c_r.alignment = Alignment(horizontal='right', vertical='center')
            _mc(ws5, r, 7, float(row.get('gasto_ads', 0) / max(row.get('conversiones', 1), 1)), fmt=EUR2)
            r += 1

        _widths(ws5, [14, 10, 14, 14, 16, 10, 14])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_reclamacion_csv(alerts: list) -> bytes:
    """Generate a clean CSV for carrier complaint submission."""
    if not alerts:
        return b''
    df = pd.DataFrame(alerts)
    df['fecha_reclamacion'] = pd.Timestamp.now().strftime('%Y-%m-%d')
    df['motivo'] = 'Coste facturado muy superior al precio contratado'
    df['accion'] = 'PENDIENTE'
    return df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
